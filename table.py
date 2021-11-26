from openpyxl import Workbook


wb = Workbook()
ws = wb.active
ws.append(["公司名称", "核准日期", "是否和最新的核准日期一致", "法人", "是否和最新的法人一致",
           "公司地址", "邮编", "是否和最新的地址一致", "是否经营异常", "网上地址", "身份证名字",
           "民族", "性别", "居住地址", "身份证号码", "签发机关", "签发日期", "失效日期", "网上核准日期"])


def company_table(rows, data: dict):
    """
    需传进公司， 核准日期， 法人， 公司地址， 邮编， 经营情况
    :param rows:
    :param data:
    :return:
    """
    ws[f"A{rows}"] = data["company_name"]
    ws[f"B{rows}"] = data["date_of_establishment"]
    ws[f"C{rows}"] = data["logic_data"]
    ws[f"D{rows}"] = data["legal_person"]
    ws[f"E{rows}"] = data["logic_person"]
    ws[f"F{rows}"] = data["address"]
    ws[f"G{rows}"] = data["post_code"]
    ws[f"H{rows}"] = data["logic_address"]
    ws[f"I{rows}"] = data["business_conditions"]
    ws[f"J{rows}"] = data["registered_address"]
    ws[f"K{rows}"] = data["name"]
    ws[f"L{rows}"] = data["ethnic"]
    ws[f"M{rows}"] = data["sex"]
    ws[f"N{rows}"] = data["living"]
    ws[f"O{rows}"] = data["citizen_id_number"]
    ws[f"P{rows}"] = data["issuing_authority"]
    ws[f"Q{rows}"] = data["date_of_issue"]
    ws[f"R{rows}"] = data["expiration_date"]
    ws[f"S{rows}"] = data["approved_date"]
    wb.save("./test.xlsx")


